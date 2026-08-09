import hashlib
import json
import re
import sys
import time
from collections import OrderedDict
from pathlib import Path

from openpyxl import load_workbook


WORKBOOK = Path(r"C:\Users\saisa\Downloads\Maths_Curriculum_Class1_to_PhD_Master.xlsx")
ASSETS_DIR = Path(__file__).resolve().parents[1] / "app" / "src" / "main" / "assets"
CONCEPT_FILE = Path(__file__).resolve().parents[1] / "app" / "src" / "main" / "java" / "com" / "indianservers" / "aiexplorer" / "MathConceptExplorer.kt"

LESSONS_JSON = ASSETS_DIR / "maths_learn_all_lessons.json"
CONCEPTS_JSON = ASSETS_DIR / "maths_concepts.json"

EXPECTED_HEADERS = [
    "Class",
    "Chapter",
    "Topic",
    "Subtopic",
    "Introduction",
    "Detailed Explanation",
    "Realtime examples",
    "Simplified Explanation",
    "Advanced Explanation",
    "Practice Prompt",
]


def clean(value):
    if value is None:
        return ""
    return str(value).replace("\r\n", "\n").strip()


def clean_header(value):
    return clean(value).rstrip(",").strip()


def slug(value):
    text = clean(value).lower()
    text = re.sub(r"[^a-z0-9]+", "-", text)
    return text.strip("-") or "item"


def kotlin(value):
    return json.dumps(value, ensure_ascii=False)


def class_rank(value):
    text = clean(value).lower()
    match = re.search(r"class\s*(\d+)", text)
    if match:
        return int(match.group(1))
    if "school enrichment" in text or "olympiad" in text:
        return 12
    if "undergraduate" in text or re.search(r"\bug\b", text):
        return 13
    if "postgraduate" in text or "msc" in text or re.search(r"\bpg\b", text):
        return 14
    if "phd" in text or "research" in text:
        return 15
    return 99


def class_band(value):
    rank = class_rank(value)
    if 1 <= rank <= 12:
        return f"Class{rank}"
    if rank == 13:
        return "UG"
    if rank == 14:
        return "PG"
    return "PhD"


def icon_for(title):
    letters = "".join(word[0] for word in re.findall(r"[A-Za-z0-9]+", title))[:3].upper()
    return letters or "M"


def summary_for(concept, topics):
    sample = ", ".join(list(topics)[:3])
    return f"{concept} lessons covering {sample}."


def concept_title_for(chapter):
    title = clean(chapter)
    compact = title.lower()
    if title in {"Number System", "Number Systems", "Numbers and Counting"}:
        return "Numbers and Number Systems"
    if title in {"Fractions", "Fractions and Decimals", "Fractions Decimals Percent", "Fractions Decimals Ratio"}:
        return "Fractions Decimals Ratio and Percent"
    if title in {"Probability and Statistics", "Statistics and Probability", "Probability and Stochastics", "Probability I", "Probability", "Statistics", "Advanced Probability", "Probability and Stochastic Processes", "Statistics and Data Science Theory", "Data and Probability"}:
        return "Probability and Statistics"
    if title in {"Data Handling", "Patterns and Data", "Data and Patterns", "Measurement and Data"}:
        return "Data Handling and Patterns"
    if title in {"Calculus", "Calculus I", "Calculus II"}:
        return "Calculus"
    if title in {"Multivariable Calculus", "Vector Calculus"}:
        return "Multivariable and Vector Calculus"
    if title in {"Differential Equations", "Partial Differential Equations", "Applied Mathematics - PDE and Scientific Computing"}:
        return "Differential Equations and PDE"
    if title in {"Linear Algebra I", "Linear Algebra II"}:
        return "Linear Algebra"
    if title in {"Abstract Algebra I", "Abstract Algebra II", "Advanced Algebra", "Pure Mathematics - Algebra", "Commutative Algebra", "Lie Theory", "Representation Theory", "Algebraic Geometry"}:
        return "Abstract and Advanced Algebra"
    if title in {"Sets Relations Functions", "Relations and Functions"}:
        return "Sets Relations and Functions"
    if title in {"Foundations", "Mathematical Logic and Foundations", "Mathematical Reasoning"}:
        return "Logic and Foundations"
    if title in {"Pure Mathematics - Analysis", "Real Analysis I", "Real Analysis II", "Advanced Real Analysis", "Functional Analysis", "Functional Analysis I", "Measure and Integration"}:
        return "Real and Functional Analysis"
    if title in {"Topology", "Topology I", "Pure Mathematics - Topology and Geometry"}:
        return "Topology and Geometry"
    if title in {"Differential Geometry", "Differential Geometry I"}:
        return "Differential Geometry"
    if title in {"Number Theory", "Pure Mathematics - Number Theory", "Algebraic Number Theory", "Analytic Number Theory", "Cryptography"}:
        return "Number Theory"
    if title in {"Optimization", "Optimization and Operations Research", "Operations Research", "Linear Programming"}:
        return "Optimization and Operations Research"
    if title in {"Discrete Mathematics", "Discrete Mathematics and Theoretical CS", "Graph Theory", "Combinatorics"}:
        return "Discrete Mathematics"
    if "dynamical systems" in compact or title == "Control Theory":
        return "Dynamical Systems and Control"
    if title in {"Commercial Mathematics", "Financial Mathematics"}:
        return "Financial and Commercial Mathematics"
    if title in {"Mathematical Modeling", "Mathematical Physics", "Mathematical Foundations of AI", "Information Theory", "Interdisciplinary Mathematics", "Mathematical Biology", "Research Practice", "Research Methods", "Problem Solving"}:
        return "Applied and Interdisciplinary Mathematics"
    if title in {"Coordinate Graphs", "Vectors and 3D Geometry"}:
        return "Coordinate Geometry"
    if title in {"Patterns and Algebra", "Patterns and Early Algebra", "Variation"}:
        return "Algebra"
    return title


def read_lessons():
    wb = load_workbook(WORKBOOK, read_only=True, data_only=True)
    ws = wb["Maths Curriculum"] if "Maths Curriculum" in wb.sheetnames else wb.active
    headers = [clean_header(cell.value) for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    if headers[: len(EXPECTED_HEADERS)] != EXPECTED_HEADERS:
        raise RuntimeError(f"Unexpected headers: {headers[:len(EXPECTED_HEADERS)]}")

    lessons = []
    seen_ids = set()
    now = int(time.time() * 1000)
    for row_index, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        values = [clean(value) for value in row[: len(EXPECTED_HEADERS)]]
        if not any(values):
            continue
        class_level, chapter, topic, subtopic, introduction, detailed, examples, simplified, advanced, practice = values
        if not (class_level and chapter and topic and subtopic):
            continue
        base_id = "__".join(slug(part) for part in (class_level, chapter, topic, subtopic))
        lesson_id = base_id
        suffix = 2
        while lesson_id in seen_ids:
            lesson_id = f"{base_id}-{suffix}"
            suffix += 1
        seen_ids.add(lesson_id)
        lessons.append(
            OrderedDict(
                [
                    ("sourceRow", row_index),
                    ("id", lesson_id),
                    ("classLevel", class_level),
                    ("chapter", chapter),
                    ("topic", topic),
                    ("subtopic", subtopic),
                    ("introduction", introduction),
                    ("detailedExplanation", detailed),
                    ("realtimeExamples", examples),
                    ("simplifiedExplanation", simplified),
                    ("advancedExplanation", advanced),
                    ("practicePrompt", practice),
                    ("updatedAt", now),
                ]
            )
        )
    return lessons


def build_concepts(lessons):
    concepts = OrderedDict()
    for lesson in lessons:
        chapter = concept_title_for(lesson["chapter"])
        entry = concepts.setdefault(
            chapter,
            {
                "title": chapter,
                "icon": icon_for(chapter),
                "topics": OrderedDict(),
                "levels": OrderedDict(),
                "lessonCount": 0,
            },
        )
        entry["lessonCount"] += 1
        entry["topics"].setdefault(lesson["topic"], 0)
        entry["topics"][lesson["topic"]] += 1
        entry["levels"].setdefault(class_band(lesson["classLevel"]), True)

    ranked = sorted(concepts.values(), key=lambda item: (-item["lessonCount"], item["title"].lower()))[:40]
    order = {name: index for index, name in enumerate([*range(1, 13), "UG", "PG", "PhD"])}
    result = []
    for entry in ranked:
        levels = sorted(entry["levels"].keys(), key=lambda name: order.get(int(name[5:]) if name.startswith("Class") else name, 99))
        topics = sorted(entry["topics"].items(), key=lambda pair: (-pair[1], pair[0].lower()))[:6]
        result.append(
            OrderedDict(
                [
                    ("title", entry["title"]),
                    ("icon", entry["icon"]),
                    ("summary", summary_for(entry["title"], [name for name, _ in topics])),
                    ("subtopics", [name for name, _ in topics]),
                    ("levels", levels),
                    ("lessonCount", entry["lessonCount"]),
                ]
            )
        )
    return sorted(result, key=lambda item: item["title"].lower())


def write_kotlin_catalog(concepts):
    concepts_code = ",\n        ".join(
        "concept({title}, {icon}, {summary}, {subs}, levels = setOf({levels}))".format(
            title=kotlin(item["title"]),
            icon=kotlin(item["icon"]),
            summary=kotlin(item["summary"]),
            subs=", ".join(kotlin(subtopic) for subtopic in item["subtopics"]),
            levels=", ".join(f"MathClassBand.{level}" for level in item["levels"]),
        )
        for item in concepts
    )
    text = CONCEPT_FILE.read_text(encoding="utf-8")
    start = text.index("    val concepts = listOf(")
    end = text.index("\n    )\n\n    fun search", start) + len("\n    )")
    replacement = f"    val concepts = listOf(\n        {concepts_code}\n    )"
    text = text[:start] + replacement + text[end:]
    text = re.sub(
        r"internal enum class MathClassBand\(val label: String\) \{.*?\n\}",
        "internal enum class MathClassBand(val label: String) {\n"
        "    Class1(\"Class 1\"), Class2(\"Class 2\"), Class3(\"Class 3\"), Class4(\"Class 4\"), Class5(\"Class 5\"),\n"
        "    Class6(\"Class 6\"), Class7(\"Class 7\"), Class8(\"Class 8\"), Class9(\"Class 9\"), Class10(\"Class 10\"),\n"
        "    Class11(\"Class 11\"), Class12(\"Class 12\"), UG(\"UG\"), PG(\"PG\"), PhD(\"PhD\"),\n"
        "}",
        text,
        flags=re.S,
    )
    text = re.sub(
        r"    private val school = .*?\n    private val senior = .*?\n    private val university = .*?\n\n",
        "",
        text,
        flags=re.S,
    )
    CONCEPT_FILE.write_text(text, encoding="utf-8")


def main():
    lessons = read_lessons()
    concepts = build_concepts(lessons)
    ASSETS_DIR.mkdir(parents=True, exist_ok=True)
    LESSONS_JSON.write_text(json.dumps(lessons, ensure_ascii=False, indent=2), encoding="utf-8")
    CONCEPTS_JSON.write_text(json.dumps(concepts, ensure_ascii=False, indent=2), encoding="utf-8")
    write_kotlin_catalog(concepts)

    digest = hashlib.sha256(LESSONS_JSON.read_bytes()).hexdigest()[:16]
    print(f"lessons={len(lessons)} concepts={len(concepts)} json_sha256={digest}")
    for concept in concepts:
        print(f"- {concept['title']} | icon={concept['icon']} | lessons={concept['lessonCount']}")


if __name__ == "__main__":
    sys.exit(main())
