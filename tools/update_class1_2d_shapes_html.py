import json
from pathlib import Path

from openpyxl import load_workbook


WORKBOOK = Path(r"C:\Users\saisa\Downloads\Maths_Curriculum_Class1_to_PhD_Master.xlsx")
LESSONS_JSON = Path(__file__).resolve().parents[1] / "app" / "src" / "main" / "assets" / "maths_learn_all_lessons.json"


def intro(shape, idea, use):
    return (
        f"<p><font color='#1565C0'><b>{shape}</b></font> is a <b>2D shape</b>. "
        f"It is flat, so we can draw it on paper.</p>"
        f"<p>{idea} Learning this helps children name shapes in pictures, toys, signs, books, and classrooms. "
        f"{use}</p>"
    )


LESSONS = {
    "Circle": {
        "Introduction": intro(
            "A circle",
            "Every point on its edge is equally far from the middle.",
            "It is useful when looking at wheels, coins, clocks, plates, and buttons.",
        ),
        "Detailed Explanation": (
            "<h3><font color='#2E7D32'>Basic idea</font></h3>"
            "<p>A <b>circle</b> is a round flat shape. It has no corners and no straight sides.</p>"
            "<h3><font color='#6A1B9A'>How to recognise it</font></h3>"
            "<ol><li>Look for a smooth round edge.</li><li>Check that there are <b>no corners</b>.</li><li>Find the middle point in your mind.</li></ol>"
            "<h3><font color='#EF6C00'>Why this is true</font></h3>"
            "<p>A circle stays round because all points on the edge are the same distance from the centre.</p>"
            "<h3><font color='#00838F'>Where we use it</font></h3>"
            "<ul><li>Wheels on a bicycle</li><li>Coins and buttons</li><li>Clock faces</li></ul>"
            "<p><b>Common mistake:</b> Calling an oval a circle.<br><b>Correct idea:</b> A circle is equally wide in every direction. An oval is longer one way.</p>"
            "<p><b>Key Point:</b> A circle is a round 2D shape with no corners.</p>"
        ),
        "Realtime examples": (
            "<ul><li><b>Easy:</b> A coin looks like a circle.</li><li><b>Compare:</b> An egg shape is not a circle because it is longer on one side.</li><li><b>Use:</b> A wheel is circular so it rolls smoothly.</li></ul>"
        ),
        "Simplified Explanation": "<p>Think of a circle like a <b>coin</b>. It is round all around. Your finger can move along its edge without stopping at any corner.</p>",
        "Advanced Explanation": "<p>In geometry, a circle is the set of all points at a fixed distance from one point called the <b>centre</b>. That fixed distance is called the <b>radius</b>.</p>",
        "Practice Prompt": "<p>Find three circle-shaped objects at home. Draw one. Then explain why it is a circle and not an oval.</p>",
    },
    "Triangle": {
        "Introduction": intro(
            "A triangle",
            "It has exactly three straight sides and three corners.",
            "It is useful when seeing roof shapes, flags, slices of pizza, and warning signs.",
        ),
        "Detailed Explanation": (
            "<h3><font color='#2E7D32'>Basic idea</font></h3>"
            "<p>A <b>triangle</b> is a flat shape made from <b>3 straight sides</b>.</p>"
            "<h3><font color='#6A1B9A'>How to recognise it</font></h3>"
            "<ol><li>Count the sides.</li><li>Check that all sides are straight.</li><li>Count the corners. A triangle has <b>3 corners</b>.</li></ol>"
            "<h3><font color='#EF6C00'>Why this is true</font></h3>"
            "<p>Three straight sides can close to make one simple flat shape. That closed shape is a triangle.</p>"
            "<h3><font color='#00838F'>Where we use it</font></h3>"
            "<ul><li>Road warning signs</li><li>Pizza slices</li><li>Strong bridge supports</li></ul>"
            "<p><b>Common mistake:</b> Calling any pointed shape a triangle.<br><b>Correct idea:</b> A triangle must have exactly three straight sides.</p>"
            "<p><b>Key Point:</b> A triangle is a closed 2D shape with 3 straight sides.</p>"
        ),
        "Realtime examples": "<ul><li><b>Easy:</b> A pizza slice can look triangular.</li><li><b>Compare:</b> A shape with 4 sides is not a triangle.</li><li><b>Use:</b> Triangles make frames strong because the sides support each other.</li></ul>",
        "Simplified Explanation": "<p>A triangle is like a <b>three-stick frame</b>. If the three sticks join end to end and close, you see a triangle.</p>",
        "Advanced Explanation": "<p>Every triangle is a polygon. A <b>polygon</b> is a closed flat shape made only from straight line segments.</p>",
        "Practice Prompt": "<p>Draw three shapes. One should be a triangle. One should have four sides. One should be open. Circle only the triangle and explain your choice.</p>",
    },
    "Square": {
        "Introduction": intro(
            "A square",
            "It has four equal sides and four square corners.",
            "It helps children recognise tiles, windows, game boards, and boxes drawn from the front.",
        ),
        "Detailed Explanation": (
            "<h3><font color='#2E7D32'>Basic idea</font></h3>"
            "<p>A <b>square</b> is a flat shape with <b>4 equal sides</b> and <b>4 right corners</b>.</p>"
            "<h3><font color='#6A1B9A'>How to recognise it</font></h3>"
            "<ol><li>Count 4 sides.</li><li>Check all sides look equal.</li><li>Check each corner looks like the corner of a book.</li></ol>"
            "<h3><font color='#EF6C00'>Why this is true</font></h3>"
            "<p>Equal sides make the shape balanced. Right corners make opposite sides line up neatly.</p>"
            "<h3><font color='#00838F'>Where we use it</font></h3>"
            "<ul><li>Floor tiles</li><li>Chessboard boxes</li><li>Sticky notes</li></ul>"
            "<p><b>Common mistake:</b> Thinking a tilted square is not a square.<br><b>Correct idea:</b> Turning a square does not change its side lengths or corners.</p>"
            "<p><b>Key Point:</b> A square has 4 equal sides and 4 right corners.</p>"
        ),
        "Realtime examples": "<ul><li><b>Easy:</b> One chessboard box is a square.</li><li><b>Compare:</b> A rectangle may have unequal long and short sides.</li><li><b>Use:</b> Square tiles fit neatly because their sides match.</li></ul>",
        "Simplified Explanation": "<p>A square is like a <b>perfect tile</b>. All four sides are the same size, and all four corners stand straight.</p>",
        "Advanced Explanation": "<p>A square is both a <b>rectangle</b> and a <b>rhombus</b>. It has right angles like a rectangle and equal sides like a rhombus.</p>",
        "Practice Prompt": "<p>Draw a square. Turn your paper sideways. Is it still a square? Explain using sides and corners.</p>",
    },
    "Rectangle": {
        "Introduction": intro(
            "A rectangle",
            "It has four straight sides and four square corners.",
            "It helps children recognise books, doors, mobile screens, boards, and tables.",
        ),
        "Detailed Explanation": (
            "<h3><font color='#2E7D32'>Basic idea</font></h3>"
            "<p>A <b>rectangle</b> is a flat shape with <b>4 sides</b> and <b>4 right corners</b>.</p>"
            "<h3><font color='#6A1B9A'>How to recognise it</font></h3>"
            "<ol><li>Count 4 sides.</li><li>Look for two long sides and two short sides.</li><li>Check that every corner is square.</li></ol>"
            "<h3><font color='#EF6C00'>Why this is true</font></h3>"
            "<p>Opposite sides of a rectangle match. Right corners keep the shape straight and neat.</p>"
            "<h3><font color='#00838F'>Where we use it</font></h3>"
            "<ul><li>Mobile phone screen</li><li>Classroom board</li><li>Book cover</li></ul>"
            "<p><b>Common mistake:</b> Thinking a square is never a rectangle.<br><b>Correct idea:</b> A square is a special rectangle because it also has 4 right corners.</p>"
            "<p><b>Key Point:</b> A rectangle has 4 sides and 4 right corners.</p>"
        ),
        "Realtime examples": "<ul><li><b>Easy:</b> A book cover is usually a rectangle.</li><li><b>Compare:</b> A circle has no straight sides, so it is not a rectangle.</li><li><b>Use:</b> Doors are rectangular so they fit straight frames.</li></ul>",
        "Simplified Explanation": "<p>A rectangle is like a <b>door shape</b>. It has two matching long sides, two matching short sides, and four neat corners.</p>",
        "Advanced Explanation": "<p>A rectangle is a quadrilateral. A <b>quadrilateral</b> is any closed flat shape with four sides.</p>",
        "Practice Prompt": "<p>Find a rectangle in your room. Count its sides and corners. Then tell whether it is also a square.</p>",
    },
    "Sorting 2D shapes": {
        "Introduction": intro(
            "Sorting 2D shapes",
            "It means grouping flat shapes by properties such as sides, corners, or curved edges.",
            "It helps children compare shapes instead of only memorising names.",
        ),
        "Detailed Explanation": (
            "<h3><font color='#2E7D32'>Basic idea</font></h3>"
            "<p><b>Sorting</b> means putting similar shapes together.</p>"
            "<h3><font color='#6A1B9A'>How to sort</font></h3>"
            "<ol><li>Choose one rule, such as number of sides.</li><li>Look at each shape carefully.</li><li>Put shapes with the same property in one group.</li></ol>"
            "<h3><font color='#EF6C00'>Why this is true</font></h3>"
            "<p>Shapes can be compared using clear properties. A property is something we can check, like <b>3 sides</b> or <b>no corners</b>.</p>"
            "<h3><font color='#00838F'>Where we use it</font></h3>"
            "<ul><li>Sorting blocks in class</li><li>Choosing tiles</li><li>Finding matching game pieces</li></ul>"
            "<p><b>Common mistake:</b> Sorting only by colour.<br><b>Correct idea:</b> In geometry, sort by shape properties first.</p>"
            "<p><b>Key Point:</b> Sort 2D shapes by clear properties like sides and corners.</p>"
        ),
        "Realtime examples": "<ul><li><b>Easy:</b> Put circles in one group and triangles in another.</li><li><b>Compare:</b> A red circle and blue circle still belong together by shape.</li><li><b>Use:</b> Builders sort tiles by shape before placing them.</li></ul>",
        "Simplified Explanation": "<p>Sorting shapes is like sorting toys. Put the ones with the same shape-rule together, even if their colours are different.</p>",
        "Advanced Explanation": "<p>Sorting introduces <b>classification</b>. Classification means grouping objects using shared properties. This prepares students for sets and geometry definitions.</p>",
        "Practice Prompt": "<p>Draw a circle, square, rectangle, and triangle. Sort them into two groups: shapes with corners and shapes without corners.</p>",
    },
    "Shape patterns": {
        "Introduction": intro(
            "A shape pattern",
            "It is a repeated order of shapes.",
            "It helps children predict what comes next in designs, tiles, games, and decorations.",
        ),
        "Detailed Explanation": (
            "<h3><font color='#2E7D32'>Basic idea</font></h3>"
            "<p>A <b>pattern</b> repeats in a planned way.</p>"
            "<h3><font color='#6A1B9A'>How to continue it</font></h3>"
            "<ol><li>Look for the smallest repeating part.</li><li>Name the order of shapes.</li><li>Repeat the same order to find the next shape.</li></ol>"
            "<h3><font color='#EF6C00'>Why this is true</font></h3>"
            "<p>The next shape follows the same rule because the pattern repeats. If the rule changes, it is not the same pattern.</p>"
            "<h3><font color='#00838F'>Where we use it</font></h3>"
            "<ul><li>Floor tiles</li><li>Gift wrapping</li><li>Classroom borders</li></ul>"
            "<p><b>Common mistake:</b> Looking only at the last shape.<br><b>Correct idea:</b> Find the repeating group first, then continue it.</p>"
            "<p><b>Key Point:</b> A shape pattern repeats the same shape order.</p>"
        ),
        "Realtime examples": "<ul><li><b>Easy:</b> Circle, square, circle, square comes next with circle.</li><li><b>Compare:</b> Circle, square, triangle, circle does not follow the same two-shape rule.</li><li><b>Use:</b> Tile designs often repeat shapes.</li></ul>",
        "Simplified Explanation": "<p>A shape pattern is like a song beat. Once you hear the repeated beat, you can tell what comes next.</p>",
        "Advanced Explanation": "<p>A repeating pattern has a <b>unit of repeat</b>. For circle, square, circle, square, the unit is circle-square.</p>",
        "Practice Prompt": "<p>Continue this pattern: circle, triangle, square, circle, triangle, square, __. Explain the repeating unit.</p>",
    },
    "Combining shapes": {
        "Introduction": intro(
            "Combining shapes",
            "It means joining simple flat shapes to make a new picture or bigger shape.",
            "It helps children see how pictures, buildings, patterns, and designs are made.",
        ),
        "Detailed Explanation": (
            "<h3><font color='#2E7D32'>Basic idea</font></h3>"
            "<p>Small <b>2D shapes</b> can join to make larger shapes or pictures.</p>"
            "<h3><font color='#6A1B9A'>How it works</font></h3>"
            "<ol><li>Choose simple shapes.</li><li>Place sides close together.</li><li>Notice the new outside shape or picture.</li></ol>"
            "<h3><font color='#EF6C00'>Why this is true</font></h3>"
            "<p>When shapes touch, their edges form a new outline. The outline tells us the new shape or picture.</p>"
            "<h3><font color='#00838F'>Where we use it</font></h3>"
            "<ul><li>Making a house from a square and triangle</li><li>Creating rangoli designs</li><li>Building pictures with paper pieces</li></ul>"
            "<p><b>Common mistake:</b> Thinking the small shapes disappear.<br><b>Correct idea:</b> The small shapes are still parts of the bigger picture.</p>"
            "<p><b>Key Point:</b> Combining shapes makes new pictures or larger shapes from smaller shapes.</p>"
        ),
        "Realtime examples": "<ul><li><b>Easy:</b> A square plus a triangle can make a house picture.</li><li><b>Compare:</b> Two triangles can make a square only when their sides fit correctly.</li><li><b>Use:</b> Artists use simple shapes to draw objects.</li></ul>",
        "Simplified Explanation": "<p>Combining shapes is like using puzzle pieces. Each piece keeps its shape, but together they make a new picture.</p>",
        "Advanced Explanation": "<p>This idea prepares students for <b>composition</b> and <b>decomposition</b>. Composition joins shapes. Decomposition breaks a shape into parts.</p>",
        "Practice Prompt": "<p>Use a square and a triangle to draw a house. Name each shape you used and point to the new outside outline.</p>",
    },
}


def update_workbook():
    wb = load_workbook(WORKBOOK)
    ws = wb["Maths Curriculum"] if "Maths Curriculum" in wb.sheetnames else wb.active
    headers = [str(cell.value).strip().rstrip(",") for cell in ws[1]]
    columns = {name: index + 1 for index, name in enumerate(headers)}
    changed = 0
    for row in range(2, ws.max_row + 1):
        class_level = ws.cell(row, columns["Class"]).value
        chapter = ws.cell(row, columns["Chapter"]).value
        topic = ws.cell(row, columns["Topic"]).value
        subtopic = ws.cell(row, columns["Subtopic"]).value
        if class_level == "Class 1" and chapter == "Geometry" and topic == "2D Shapes" and subtopic in LESSONS:
            lesson = LESSONS[subtopic]
            for field, value in lesson.items():
                ws.cell(row, columns[field]).value = value
            changed += 1
    wb.save(WORKBOOK)
    return changed


def update_json():
    lessons = json.loads(LESSONS_JSON.read_text(encoding="utf-8"))
    changed = 0
    for lesson in lessons:
        if lesson["classLevel"] == "Class 1" and lesson["chapter"] == "Geometry" and lesson["topic"] == "2D Shapes" and lesson["subtopic"] in LESSONS:
            values = LESSONS[lesson["subtopic"]]
            lesson["introduction"] = values["Introduction"]
            lesson["detailedExplanation"] = values["Detailed Explanation"]
            lesson["realtimeExamples"] = values["Realtime examples"]
            lesson["simplifiedExplanation"] = values["Simplified Explanation"]
            lesson["advancedExplanation"] = values["Advanced Explanation"]
            lesson["practicePrompt"] = values["Practice Prompt"]
            changed += 1
    LESSONS_JSON.write_text(json.dumps(lessons, ensure_ascii=False, indent=2), encoding="utf-8")
    return changed


if __name__ == "__main__":
    print(f"workbook_rows={update_workbook()}")
    print(f"json_rows={update_json()}")
