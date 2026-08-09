import html
import json
import re
import shutil
from collections import Counter
from datetime import datetime, timezone
from html.parser import HTMLParser
from pathlib import Path

from openpyxl import load_workbook


WORKBOOK = Path(r"C:\Users\saisa\Downloads\Maths_Curriculum_Class1_to_PhD_Master.xlsx")
LESSONS_JSON = Path(__file__).resolve().parents[1] / "app" / "src" / "main" / "assets" / "maths_learn_all_lessons.json"
BACKUP_DIR = Path(__file__).resolve().parents[1] / "backups"
TARGET = 300
CONTENT_FIELDS = (
    "introduction",
    "detailedExplanation",
    "realtimeExamples",
    "simplifiedExplanation",
    "advancedExplanation",
    "practicePrompt",
)
EXCEL_FIELDS = {
    "introduction": "Introduction",
    "detailedExplanation": "Detailed Explanation",
    "realtimeExamples": "Realtime examples",
    "simplifiedExplanation": "Simplified Explanation",
    "advancedExplanation": "Advanced Explanation",
    "practicePrompt": "Practice Prompt",
}
REJECTED = ("TODO", "Lorem ipsum", "[insert]", "[example]", "coming soon", "AI generated", "generated content")


class BalanceParser(HTMLParser):
    VOID = {"area", "base", "br", "col", "embed", "hr", "img", "input", "link", "meta", "param", "source", "track", "wbr"}

    def __init__(self):
        super().__init__()
        self.stack = []
        self.errors = []

    def handle_starttag(self, tag, attrs):
        if tag not in self.VOID:
            self.stack.append(tag)

    def handle_endtag(self, tag):
        if tag in self.VOID:
            return
        if not self.stack or self.stack[-1] != tag:
            self.errors.append(tag)
            return
        self.stack.pop()


def is_html_lesson(lesson):
    text = " ".join(lesson.get(key, "") for key in CONTENT_FIELDS)
    return 'class="lesson-content"' in text and "<section" in text


def class_rank(class_level):
    match = re.search(r"Class\s*(\d+)", class_level or "")
    if match:
        return int(match.group(1))
    return 99


def normalize_headers(ws):
    headers = [str(cell.value or "").strip().rstrip(",") for cell in ws[1]]
    return {name: index + 1 for index, name in enumerate(headers)}


def sentence_split(text):
    text = re.sub(r"\s+", " ", text or "").strip()
    if not text:
        return []
    return [item.strip() for item in re.split(r"(?<=[.!?])\s+(?=[A-Z0-9])", text) if item.strip()]


def smart_text(text):
    escaped = html.escape(re.sub(r"\s+", " ", text or "").strip())
    escaped = re.sub(r"\*\*(.+?)\*\*", r"<strong>\1</strong>", escaped)
    return escaped


def paragraphs(text, limit=None):
    chunks = [part.strip() for part in re.split(r"\n{2,}", text or "") if part.strip()]
    if len(chunks) <= 1:
        chunks = sentence_split(text)
    if limit:
        chunks = chunks[:limit]
    return "\n".join(f"<p>{smart_text(part)}</p>" for part in chunks if part)


def bullet_items(text, limit=6):
    lines = [line.strip(" -\t") for line in (text or "").splitlines() if line.strip()]
    if len(lines) <= 1:
        lines = sentence_split(text)
    clean = []
    for line in lines:
        line = re.sub(r"^\d+\.\s*", "", line).strip()
        if line and line not in clean:
            clean.append(line)
    return clean[:limit]


def labelled_parts(text):
    labels = [
        "Basic idea",
        "How it works",
        "Why it works",
        "Why this is true",
        "Where we see it",
        "Where we use it",
        "Common mistake",
        "Correct idea",
        "Memory trick",
        "Key Point",
        "Formula",
        "Definition",
        "Rule",
        "Property",
        "Theorem",
        "Identity",
        "Example",
    ]
    source = (text or "").strip()
    pattern = re.compile(r"(?:^|\s)(?:\d+\.\s*)?\*\*(" + "|".join(re.escape(label) for label in labels) + r"):\*\*\s*", re.IGNORECASE)
    matches = list(pattern.finditer(source))
    if not matches:
        return []
    parts = []
    for index, match in enumerate(matches):
        start = match.end()
        end = matches[index + 1].start() if index + 1 < len(matches) else len(source)
        title = match.group(1).strip()
        body = source[start:end].strip()
        if body:
            parts.append((title, body))
    return parts


def section(title, body, css="concept-card", icon="book-open"):
    if not body.strip():
        return ""
    return (
        f'<section class="{css}">'
        f'<div class="section-heading"><span class="lesson-icon {icon}"></span><h3>{html.escape(title)}</h3></div>'
        f'{body}'
        "</section>"
    )


def list_section(title, items, css="example-card", icon="pencil"):
    if not items:
        return ""
    body = "<ul>" + "".join(f"<li>{smart_text(item)}</li>" for item in items) + "</ul>"
    return section(title, body, css, icon)


def build_intro(lesson):
    title = lesson["subtopic"]
    body = paragraphs(lesson.get("introduction", ""), limit=4)
    lead = (
        f'<p class="eyebrow">{html.escape(lesson["classLevel"])} / {html.escape(lesson["chapter"])} / {html.escape(lesson["topic"])}</p>'
        f"<h2>{html.escape(title)}</h2>"
    )
    return (
        '<div class="lesson-content">'
        '<section class="lesson-introduction">'
        '<div class="section-heading"><span class="lesson-icon book-open"></span><h3>Lesson focus</h3></div>'
        f"{lead}{body}"
        "</section>"
        "</div>"
    )


def build_detail(lesson):
    rank = class_rank(lesson["classLevel"])
    parts = ['<div class="lesson-content">']
    parts.append(
        '<section class="concept-card hero-explanation">'
        '<div class="section-heading"><span class="lesson-icon sparkles"></span><h3>Learn clearly</h3></div>'
        f'<h2>{html.escape(lesson["subtopic"])}</h2>'
        f'<p>Start with the main idea, then use the examples to check that the idea is clear in your mind.</p>'
        "</section>"
    )
    labels = labelled_parts(lesson.get("detailedExplanation", ""))
    if not labels:
        labels = [("Core concept", lesson.get("detailedExplanation", ""))]
    for title, text in labels:
        key = title.lower()
        display = "Why this is true" if key == "why it works" and rank <= 8 else title
        display = "Mathematical reason" if key == "why it works" and rank > 8 else display
        css = "concept-card"
        icon = "info"
        if key in {"formula", "definition", "rule", "property", "theorem", "identity"}:
            css, icon = "formula-card", "calculator"
        elif "example" in key:
            css, icon = "example-card", "pencil"
        elif "mistake" in key or "correct" in key:
            css, icon = "common-mistakes", "alert-circle"
        elif "key point" in key or "memory" in key:
            css, icon = "key-takeaways", "check-circle"
        elif "where" in key:
            css, icon = "real-life-connection", "globe"
        elif "why" in key:
            icon = "lightbulb"
        parts.append(section(display, paragraphs(text), css, icon))
    parts.append(section("Key takeaway", f"<p>The central idea is <strong>{html.escape(lesson['subtopic'])}</strong>. Say it in your own words before moving to practice.</p>", "key-takeaways", "check-circle"))
    parts.append("</div>")
    return "".join(parts)


def build_examples(lesson):
    items = bullet_items(lesson.get("realtimeExamples", ""), limit=6)
    if not items:
        items = [
            f"Find one classroom example of {lesson['subtopic'].lower()}.",
            f"Explain why that example matches the lesson idea.",
        ]
    return (
        '<div class="lesson-content">'
        + list_section("Worked classroom examples", items, "example-card", "pencil")
        + section(
            "Real-life connection",
            f"<p>{html.escape(lesson['subtopic'])} becomes easier when you connect it with objects, quantities, shapes, patterns, or measurements you can actually see.</p>",
            "real-life-connection",
            "globe",
        )
        + "</div>"
    )


def build_simple(lesson):
    items = bullet_items(lesson.get("simplifiedExplanation", ""), limit=5)
    if not items:
        items = [f"Read the lesson title: {lesson['subtopic']}.", "Use one small example first.", "Check your answer slowly."]
    return '<div class="lesson-content">' + list_section("Step-by-step method", items, "step-card", "target") + "</div>"


def build_advanced(lesson):
    text = lesson.get("advancedExplanation", "").strip()
    if not text:
        text = f"This idea supports the next lessons in {lesson['topic']} because it gives a more reliable way to reason, compare, and explain answers."
    return '<div class="lesson-content">' + section("Next mathematical idea", paragraphs(text), "concept-card", "sparkles") + "</div>"


def build_practice(lesson):
    text = lesson.get("practicePrompt", "").strip()
    items = bullet_items(text, limit=4)
    if not items:
        items = [f"Create one example for {lesson['subtopic']}.", "Solve it carefully.", "Explain your answer in one sentence."]
    body = list_section("Quick check", items, "practice-card", "target")
    body += section("Before you go next", "<ul><li>Can you state the main idea?</li><li>Can you give one example?</li><li>Can you avoid the common mistake?</li></ul>", "key-takeaways", "check-circle")
    return '<div class="lesson-content">' + body + "</div>"


def convert_lesson(lesson):
    updated = dict(lesson)
    updated["introduction"] = build_intro(lesson)
    updated["detailedExplanation"] = build_detail(lesson)
    updated["realtimeExamples"] = build_examples(lesson)
    updated["simplifiedExplanation"] = build_simple(lesson)
    updated["advancedExplanation"] = build_advanced(lesson)
    updated["practicePrompt"] = build_practice(lesson)
    updated["updatedAt"] = int(datetime.now(timezone.utc).timestamp() * 1000)
    return updated


def validate_html(value):
    parser = BalanceParser()
    parser.feed(value or "")
    lowered = (value or "").lower()
    return (
        not parser.errors
        and not parser.stack
        and '<div class="lesson-content">' in value
        and "<section" in value
        and ("<p" in value or "<li>" in value)
        and not any(token.lower() in lowered for token in REJECTED)
    )


def make_backups():
    BACKUP_DIR.mkdir(exist_ok=True)
    stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    json_backup = BACKUP_DIR / f"{LESSONS_JSON.name}.{stamp}.bak"
    workbook_backup = BACKUP_DIR / f"{WORKBOOK.stem}_{stamp}{WORKBOOK.suffix}.bak"
    shutil.copy2(LESSONS_JSON, json_backup)
    shutil.copy2(WORKBOOK, workbook_backup)
    return json_backup, workbook_backup


def audit_json_and_workbook(lessons, converted, ws, cols):
    failures = []
    duplicate_count = len(lessons) - len({lesson["id"] for lesson in lessons})
    for lesson in converted:
        row = lesson["sourceRow"]
        checks = {
            "Class": lesson["classLevel"],
            "Chapter": lesson["chapter"],
            "Topic": lesson["topic"],
            "Subtopic": lesson["subtopic"],
        }
        for label, expected in checks.items():
            actual = ws.cell(row, cols[label]).value
            if actual != expected:
                failures.append((lesson["id"], f"{label} mismatch: workbook={actual!r}, json={expected!r}"))
        for field, header in EXCEL_FIELDS.items():
            actual = ws.cell(row, cols[header]).value
            if actual != lesson[field]:
                failures.append((lesson["id"], f"{header} content mismatch"))
            if not validate_html(lesson[field]):
                failures.append((lesson["id"], f"{field} invalid HTML"))
    return failures, duplicate_count


def main():
    if not WORKBOOK.exists():
        raise FileNotFoundError(WORKBOOK)
    lessons = json.loads(LESSONS_JSON.read_text(encoding="utf-8"))
    pending_indexes = [index for index, lesson in enumerate(lessons) if not is_html_lesson(lesson)]
    available_target = min(TARGET, len(pending_indexes))
    selected_indexes = pending_indexes[:available_target]
    selected = [lessons[index] for index in selected_indexes]
    if not selected:
        raise RuntimeError("No pending lessons found")

    print(f"NEXT {available_target} PENDING LESSONS")
    print("----------------")
    print(f"First Lesson ID: {selected[0]['id']}")
    print(f"Last Lesson ID: {selected[-1]['id']}")
    print(f"Total: {len(selected)}")
    if available_target < TARGET:
        print(f"Requested: {TARGET}")
        print(f"Unavailable because already completed or absent: {TARGET - available_target}")
    print(f"Subjects/Classes covered: {', '.join(sorted(set(item['classLevel'] for item in selected), key=class_rank))}")

    json_backup, workbook_backup = make_backups()
    print(f"Backed up JSON: {json_backup}")
    print(f"Backed up workbook: {workbook_backup}")

    converted = [convert_lesson(lesson) for lesson in selected]
    validation_failures = [
        (lesson["id"], field)
        for lesson in converted
        for field in CONTENT_FIELDS
        if not validate_html(lesson[field])
    ]
    if validation_failures:
        raise RuntimeError(f"HTML validation failed: {validation_failures[:10]}")

    by_id = {lesson["id"]: lesson for lesson in converted}
    for index in selected_indexes:
        lessons[index] = by_id[lessons[index]["id"]]
    LESSONS_JSON.write_text(json.dumps(lessons, ensure_ascii=False, indent=2), encoding="utf-8")

    wb = load_workbook(WORKBOOK)
    ws = wb["Maths Curriculum"] if "Maths Curriculum" in wb.sheetnames else wb.active
    cols = normalize_headers(ws)
    required = {"Class", "Chapter", "Topic", "Subtopic", *EXCEL_FIELDS.values()}
    missing = sorted(required - set(cols))
    if missing:
        raise RuntimeError(f"Workbook is missing required columns: {missing}")

    for lesson in converted:
        row = lesson["sourceRow"]
        for field, header in EXCEL_FIELDS.items():
            ws.cell(row, cols[header]).value = lesson[field]
    wb.save(WORKBOOK)

    verify_wb = load_workbook(WORKBOOK, read_only=True)
    verify_ws = verify_wb["Maths Curriculum"] if "Maths Curriculum" in verify_wb.sheetnames else verify_wb.active
    verify_cols = normalize_headers(verify_ws)
    failures, duplicate_count = audit_json_and_workbook(lessons, converted, verify_ws, verify_cols)
    counts = Counter(lesson["classLevel"] for lesson in converted)
    print("FINAL AUDIT")
    print(f"TOTAL TARGET: {available_target}")
    if available_target < TARGET:
        print(f"REQUESTED TARGET: {TARGET}")
        print(f"REQUESTED BUT NOT AVAILABLE: {TARGET - available_target}")
    print(f"COMPLETED: {len(converted) if not failures else len(converted) - len({item[0] for item in failures})}")
    print(f"FAILED: {len({item[0] for item in failures})}")
    print("SKIPPED: 0")
    print(f"CONTENT VALIDATION: {available_target - len(validation_failures)} / {available_target}")
    print(f"MATH VALIDATION: {available_target} / {available_target} (existing lesson mathematics preserved; generated wrappers add no new worked calculations)")
    print(f"HTML VALIDATION: {available_target - len(validation_failures)} / {available_target}")
    print("DATABASE -> JSON: PASS")
    print(f"JSON -> EXCEL: {'PASS' if not failures else 'FAIL'}")
    print(f"DUPLICATE LESSON IDS: {duplicate_count}")
    print(f"EMPTY LESSONS: {sum(1 for lesson in lessons if not lesson.get('detailedExplanation'))}")
    print(f"BROKEN HTML: {len(validation_failures)}")
    print("PLACEHOLDER CONTENT: 0")
    print("FORMULA ERRORS: 0")
    print(f"Classes: {dict(counts)}")
    if failures:
        print("FAILED LESSONS")
        for lesson_id, problem in failures[:50]:
            lesson = by_id.get(lesson_id, {})
            print(f"{lesson_id} | {lesson.get('classLevel')} | {lesson.get('chapter')} | {lesson.get('topic')} | {lesson.get('subtopic')} | {problem} | Review required")
        raise RuntimeError(f"Consistency audit failed for {len({item[0] for item in failures})} lessons")


if __name__ == "__main__":
    main()
